"""
Endpoints de administração
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session
from io import BytesIO
from app.core.database import get_db
from app.core.dependencies import get_current_admin, get_current_user
from app.core.security import get_password_hash
from app.models.user import User, UserType
from app.models.job import Job
from app.models.job_application import JobApplication
from app.models.candidate import Candidate
from app.models.company import Company
from app.models.test import Test, Question, Alternative, TestLevel
from app.models.competencia import Competencia, AreaAtuacao
from app.schemas.test import TestCreate, TestUpdate, TestResponse, TestListResponse, TestListItemResponse, QuestionListItem, TestCreateRequest
from app.schemas.competencia import CompetenciaCreate, CompetenciaResponse
from app.schemas.job import JobCreate
from app.services.test_import_service import TestImportService
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional
import logging

# Setup de logs
logger = logging.getLogger(__name__)

router = APIRouter(tags=["admin"])


class CreateadminRequest(BaseModel):
    """Schema para criação de admin"""
    email: EmailStr
    password: str = Field(..., min_length=6)


@router.get("/debug/usuarios-vs-empresas")
async def debug_usuarios_vs_empresas(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Endpoint de debug para verificar discrepância entre usuários e empresas
    """
    # Contar usuários do tipo empresa
    usuarios_empresa = db.query(User).filter(User.user_type == UserType.empresa).all()
    
    # Contar empresas cadastradas
    empresas = db.query(Company).all()
    
    # Verificar quais usuários empresa não têm empresa vinculada
    usuarios_sem_empresa = []
    for user in usuarios_empresa:
        company = db.query(Company).filter(Company.user_id == user.id).first()
        if not company:
            usuarios_sem_empresa.append({
                "user_id": user.id,
                "email": user.email,
                "created_at": user.created_at
            })
    
    return {
        "total_usuarios_empresa": len(usuarios_empresa),
        "total_empresas_cadastradas": len(empresas),
        "usuarios_sem_empresa_vinculada": len(usuarios_sem_empresa),
        "detalhes_usuarios_sem_empresa": usuarios_sem_empresa,
        "empresas": [
            {
                "id": c.id,
                "cnpj": c.cnpj,
                "razao_social": c.razao_social,
                "user_id": c.user_id
            }
            for c in empresas
        ]
    }



@router.get("/debug/companies-raw")
async def debug_companies_raw(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna TODOS os dados brutos da tabela companies diretamente do banco
    """
    empresas = db.query(Company).all()
    
    print(f"DEBUG /debug/companies-raw: Total de empresas no banco: {len(empresas)}")
    
    return {
        "total_empresas": len(empresas),
        "empresas": [
            {
                "id": c.id,
                "cnpj": c.cnpj,
                "razao_social": c.razao_social,
                "nome_fantasia": c.nome_fantasia,
                "setor": c.setor,
                "user_id": c.user_id,
                "is_active": c.is_active,
                "is_verified": c.is_verified,
                "created_at": c.created_at,
                "updated_at": c.updated_at
            }
            for c in empresas
        ]
    }


@router.post("/fix/criar-empresa-faltante")
async def criar_empresa_faltante(
    user_id: int,
    cnpj: str = "",
    razao_social: str = "Empresa Padrão",
    nome_fantasia: str = "Empresa",
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    **ENDPOINT DE CORREÇÃO**: Cria uma empresa para um usuário que não tem
    
    Use quando um usuário do tipo 'empresa' não tiver registro na tabela companies
    
    Parâmetros:
    - user_id: ID do usuário (tipo empresa)
    - cnpj: CNPJ da empresa (opcional)
    - razao_social: Razão social (padrão: "Empresa Padrão")
    - nome_fantasia: Nome fantasia (padrão: "Empresa")
    
    Exemplo:
    POST /api/v1/admin/fix/criar-empresa-faltante?user_id=16&cnpj=12345678000195&razao_social=MinhaEmpresa
    """
    try:
        # Verificar se o usuário existe
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Usuário {user_id} não encontrado"
            )
        
        # Verificar se é do tipo empresa
        if user.user_type != UserType.empresa:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Usuário {user_id} não é do tipo 'empresa' (type={user.user_type.value})"
            )
        
        # Verificar se já tem empresa
        existing = db.query(Company).filter(Company.user_id == user_id).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Usuário {user_id} já tem uma empresa cadastrada (company_id={existing.id})"
            )
        
        # Criar empresa
        new_company = Company(
            user_id=user_id,
            cnpj=cnpj or f"00000000000{user_id:03d}",  # CNPJ fallback
            razao_social=razao_social,
            nome_fantasia=nome_fantasia,
            is_active=True,
            is_verified=True
        )
        
        db.add(new_company)
        db.commit()
        db.refresh(new_company)
        
        logger.info(f"Empresa criada: id={new_company.id} para user_id={user_id}")
        
        return {
            "mensagem": "Empresa criada com sucesso",
            "company_id": new_company.id,
            "user_id": new_company.user_id,
            "cnpj": new_company.cnpj,
            "razao_social": new_company.razao_social,
            "nome_fantasia": new_company.nome_fantasia
        }
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Erro ao criar empresa: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao criar empresa: {str(e)}"
        )


@router.get("/debug/auth")
async def debug_auth(
    current_user: User = Depends(get_current_user)
):
    """
    Endpoint de debug para verificar autenticação
    Retorna informações do usuário autenticado
    """
    return {
        "status": "autenticado",
        "user_id": current_user.id,
        "email": current_user.email,
        "user_type": current_user.user_type.value,
        "is_active": current_user.is_active,
        "is_verified": current_user.is_verified
    }


class CreateadminRequest(BaseModel):
    """Schema para criação de admin"""
    email: EmailStr
    password: str = Field(..., min_length=6)


@router.get("/dashboard/stats")
async def get_admin_dashboard_stats(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna estatísticas gerais do painel administrativo
    
    Retorna:
    - Total de candidatos
    - Candidatos cadastrados nos últimos 30 dias
    - Total de empresas
    - Total de vagas abertas
    - Total de candidaturas
    """
    from datetime import datetime, timedelta
    
    total_candidatos = db.query(Candidate).count()
    
    # Candidatos dos últimos 30 dias
    trinta_dias_atras = datetime.utcnow() - timedelta(days=30)
    candidatos_30_dias = db.query(Candidate).filter(
        Candidate.created_at >= trinta_dias_atras
    ).count()
    
    total_empresas = db.query(Company).count()
    
    # Import JobStatus from models
    from app.models.job import JobStatus
    total_vagas_abertas = db.query(Job).filter(
        (Job.status == JobStatus.ABERTA) | (Job.status == JobStatus.PAUSADA)
    ).count()
    total_candidaturas = db.query(JobApplication).count()
    
    return {
        "total_candidatos": total_candidatos,
        "candidatos_ultimos_30_dias": candidatos_30_dias,
        "total_empresas": total_empresas,
        "total_vagas_abertas": total_vagas_abertas,
        "total_candidaturas": total_candidaturas
    }


@router.get("/usuarios")
async def get_all_users(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(10, ge=1, le=100, description="Número máximo de registros a retornar (máximo 100)"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todos os usuários registrados no sistema com paginação
    
    Apenas administradores podem acessar este endpoint.
    
    Query Parameters:
    - skip: Número de registros a pular (default: 0)
    - limit: Número máximo de registros por página (default: 10, máximo: 100)
    """
    total = db.query(User).count()
    users = db.query(User).offset(skip).limit(limit).all()
    
    # Formatar resposta com informações relevantes
    users_list = []
    for user in users:
        user_info = {
            "id": user.id,
            "email": user.email,
            "user_type": user.user_type.value,
            "is_active": user.is_active,
            "is_verified": user.is_verified,
            "created_at": user.created_at,
            "updated_at": user.updated_at
        }
        
        # Adicionar informações específicas baseado no tipo de usuário
        if user.user_type == UserType.empresa and user.company:
            user_info["company"] = {
                "cnpj": user.company.cnpj,
                "razao_social": user.company.razao_social,
                "setor": user.company.setor,
                "is_active": user.company.is_active
            }
        elif user.user_type == UserType.candidato and user.candidate:
            user_info["candidate"] = {
                "cpf": user.candidate.cpf,
                "full_name": user.candidate.full_name,
                "phone": user.candidate.phone
            }
        
        users_list.append(user_info)
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "returned": len(users_list),
        "usuarios": users_list
    }


@router.get("/vagas")
async def get_all_jobs(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(10, ge=1, le=100, description="Número máximo de registros a retornar (máximo 100)"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todas as vagas cadastradas no sistema com paginação
    
    Apenas administradores podem acessar este endpoint.
    
    Query Parameters:
    - skip: Número de registros a pular (default: 0)
    - limit: Número máximo de registros por página (default: 10, máximo: 100)
    """
    total = db.query(Job).count()
    jobs = db.query(Job).offset(skip).limit(limit).all()
    
    # Formatar resposta com informações relevantes
    jobs_list = []
    for job in jobs:
        job_info = {
            "id": job.id,
            "title": job.title,
            "location": job.location,
            "remote": job.remote,
            "job_type": job.job_type,
            "status": job.status.value,
            "salary_min": float(job.salary_min) if job.salary_min else None,
            "salary_max": float(job.salary_max) if job.salary_max else None,
            "salary_currency": job.salary_currency,
            "views_count": job.views_count,
            "applications_count": job.applications_count,
            "created_at": job.created_at,
            "updated_at": job.updated_at,
            "published_at": job.published_at,
            "closed_at": job.closed_at,
            "company": {
                "cnpj": job.company.cnpj,
                "razao_social": job.company.razao_social
            } if job.company else None
        }
        jobs_list.append(job_info)
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "returned": len(jobs_list),
        "vagas": jobs_list
    }


@router.post("/vagas", response_model=dict, status_code=status.HTTP_201_CREATED)
async def create_job_admin(
    job_data: JobCreate,
    company_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Cria uma nova vaga para qualquer empresa (apenas admin)
    
    Apenas administradores podem acessar este endpoint.
    Requer company_id como query parameter para indicar qual empresa a vaga pertence.
    
    Exemplo de payload:
    {
        "company_id": 1,
        "title": "Desenvolvedor Python",
        "description": "Descrição da vaga",
        "requirements": "Python, FastAPI",
        "location": "São Paulo",
        "job_type": "full-time",
        "salary_min": 5000,
        "salary_max": 8000,
        "salary_currency": "BRL",
        "remote": false,
        "benefits": "Vale refeição, vale transporte",
        "status": "rascunho"
    }
    """
    # Verificar se a empresa existe
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Empresa com ID {company_id} não encontrada"
        )
    
    try:
        print(f"DEBUG - Admin criando vaga para empresa {company_id}")
        print(f"DEBUG - Dados: title={job_data.title}, location={job_data.location}")
        
        from app.services.job_service import JobService
        service = JobService(db)
        job = await service.create_job(company_id, job_data)
        
        print(f"DEBUG - Vaga criada com sucesso: ID={job.id}")
        
        return {
            "id": job.id,
            "title": job.title,
            "company_id": job.company_id,
            "status": job.status.value,
            "created_at": job.created_at,
            "message": "Vaga criada com sucesso"
        }
    except Exception as e:
        print(f"ERRO ao criar vaga: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao criar vaga: {str(e)}"
        )


@router.get("/vagas/{job_id}")
async def get_job_admin(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna detalhes de uma vaga específica (qualquer admin pode ver)
    
    Apenas administradores podem acessar este endpoint.
    """
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    return {
        "id": job.id,
        "title": job.title,
        "description": job.description,
        "location": job.location,
        "remote": job.remote,
        "job_type": job.job_type,
        "status": job.status.value,
        "salary_min": float(job.salary_min) if job.salary_min else None,
        "salary_max": float(job.salary_max) if job.salary_max else None,
        "salary_currency": job.salary_currency,
        "views_count": job.views_count,
        "applications_count": job.applications_count,
        "created_at": job.created_at,
        "updated_at": job.updated_at,
        "published_at": job.published_at,
        "closed_at": job.closed_at,
        "company": {
            "id": job.company.id,
            "cnpj": job.company.cnpj,
            "razao_social": job.company.razao_social
        } if job.company else None
    }


@router.put("/vagas/{job_id}")
async def update_job_admin(
    job_id: int,
    job_update: dict,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Atualiza uma vaga (qualquer admin pode editar)
    
    Apenas administradores podem acessar este endpoint.
    """
    from app.schemas.job import JobUpdate
    
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    try:
        # Atualizar apenas os campos fornecidos
        for field, value in job_update.items():
            if hasattr(job, field) and value is not None:
                setattr(job, field, value)
        
        db.commit()
        db.refresh(job)
        
        return {
            "id": job.id,
            "title": job.title,
            "status": job.status.value,
            "updated_at": job.updated_at,
            "message": "Vaga atualizada com sucesso"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao atualizar vaga: {str(e)}"
        )


@router.delete("/vagas/{job_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_job_admin(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Deleta uma vaga (qualquer admin pode deletar)
    
    Apenas administradores podem acessar este endpoint.
    Remove a vaga e todas as suas candidaturas.
    """
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    try:
        db.delete(job)
        db.commit()
        return None
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao deletar vaga: {str(e)}"
        )


@router.post("/vagas/{job_id}/publish")
async def publish_job_admin(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Publica uma vaga (qualquer admin pode publicar)
    
    Apenas administradores podem acessar este endpoint.
    """
    from app.models.job import JobStatus
    from datetime import datetime
    
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    try:
        job.status = JobStatus.ABERTA
        job.published_at = datetime.utcnow()
        db.commit()
        db.refresh(job)
        
        return {
            "id": job.id,
            "title": job.title,
            "status": job.status.value,
            "published_at": job.published_at,
            "message": "Vaga publicada com sucesso"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao publicar vaga: {str(e)}"
        )


@router.post("/vagas/{job_id}/close")
async def close_job_admin(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Encerra uma vaga (qualquer admin pode encerrar)
    
    Apenas administradores podem acessar este endpoint.
    """
    from app.models.job import JobStatus
    from datetime import datetime
    
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    try:
        job.status = JobStatus.ENCERRADA
        job.closed_at = datetime.utcnow()
        db.commit()
        db.refresh(job)
        
        return {
            "id": job.id,
            "title": job.title,
            "status": job.status.value,
            "closed_at": job.closed_at,
            "message": "Vaga encerrada com sucesso"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao encerrar vaga: {str(e)}"
        )


@router.get("/candidaturas")
async def get_all_applications(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(10, ge=1, le=100, description="Número máximo de registros a retornar (máximo 100)"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todas as candidaturas de todas as vagas com paginação
    
    Apenas administradores podem acessar este endpoint.
    
    Query Parameters:
    - skip: Número de registros a pular (default: 0)
    - limit: Número máximo de registros por página (default: 10, máximo: 100)
    """
    total = db.query(JobApplication).count()
    applications = db.query(JobApplication).offset(skip).limit(limit).all()
    
    # Formatar resposta com informações relevantes
    applications_list = []
    for app in applications:
        app_info = {
            "id": app.id,
            "status": app.status.value,
            "created_at": app.created_at,
            "updated_at": app.updated_at,
            "job": {
                "id": app.job.id,
                "title": app.job.title,
                "company": {
                    "cnpj": app.job.company.cnpj,
                    "razao_social": app.job.company.razao_social
                }
            } if app.job else None,
            "candidate": {
                "cpf": app.candidate.cpf,
                "full_name": app.candidate.full_name,
                "email": app.candidate.user.email if app.candidate.user else None
            } if app.candidate else None
        }
        applications_list.append(app_info)
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "returned": len(applications_list),
        "candidaturas": applications_list
    }


@router.get("/vagas/{vaga_id}/candidaturas")
async def get_vaga_candidaturas(
    vaga_id: int,
    status: Optional[str] = Query(None, description="Filtrar por status"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todas as candidaturas de uma vaga específica
    
    Apenas administradores podem acessar este endpoint.
    
    Parâmetros:
    - vaga_id: ID da vaga
    - status: Filtro por status (em_analise, entrevista, finalista, recusado, contratado)
    """
    # Verificar se a vaga existe
    job = db.query(Job).filter(Job.id == vaga_id).first()
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vaga não encontrada"
        )
    
    # Buscar candidaturas
    query = db.query(JobApplication).filter(JobApplication.job_id == vaga_id)
    
    if status:
        query = query.filter(JobApplication.status == status)
    
    applications = query.all()
    
    # Formatar resposta
    applications_list = []
    for app in applications:
        app_info = {
            "id": app.id,
            "status": app.status.value,
            "created_at": app.created_at,
            "updated_at": app.updated_at,
            "candidate": {
                "id": app.candidate.id,
                "cpf": app.candidate.cpf,
                "nome": app.candidate.full_name,
                "email": app.candidate.user.email if app.candidate.user else None,
                "telefone": app.candidate.phone,
                "localizacao": app.candidate.estado,
                "habilidades": app.candidate.habilidades
            } if app.candidate else None
        }
        applications_list.append(app_info)
    
    return {
        "vaga_id": vaga_id,
        "vaga_titulo": job.title,
        "total": len(applications_list),
        "candidaturas": applications_list
    }


@router.post("/create-admin")
async def create_admin(
    admin_data: CreateadminRequest,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Cria um novo administrador

    Apenas administradores podem criar outros administradores.
    """
    existing_user = db.query(User).filter(User.email == admin_data.email).first()
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email já cadastrado no sistema"
        )

    new_admin = User(
        email=admin_data.email,
        password_hash=get_password_hash(admin_data.password),
        user_type=UserType.admin,
        is_active=True,
        is_verified=True
    )

    db.add(new_admin)
    db.commit()
    db.refresh(new_admin)

    return {
        "id": new_admin.id,
        "email": new_admin.email,
        "user_type": new_admin.user_type.value,
        "is_active": new_admin.is_active,
        "created_at": new_admin.created_at,
        "message": "administrador criado com sucesso"
    }


class CreateCompanyForUserRequest(BaseModel):
    """Schema para criar empresa para um usuário existente"""
    user_id: int
    cnpj: str
    razao_social: str
    nome_fantasia: str = None
    setor: str = None
    cep: str = None
    pessoa_de_contato: str = None
    fone: str = None
    site: str = None


@router.post("/criar-empresa-usuario")
async def create_company_for_user(
    company_data: CreateCompanyForUserRequest,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Cria uma empresa para um usuário existente que não tem empresa vinculada
    
    Apenas administradores podem acessar este endpoint.
    """
    # Verificar se usuário existe e é do tipo empresa
    user = db.query(User).filter(User.id == company_data.user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuário não encontrado"
        )
    
    if user.user_type != UserType.empresa:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Usuário não é do tipo empresa"
        )
    
    # Verificar se usuário já tem empresa
    existing_company = db.query(Company).filter(Company.user_id == company_data.user_id).first()
    if existing_company:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Usuário já possui empresa cadastrada"
        )
    
    # Verificar se CNPJ já existe
    existing_cnpj = db.query(Company).filter(Company.cnpj == company_data.cnpj).first()
    if existing_cnpj:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CNPJ já cadastrado"
        )
    
    # Criar empresa
    company = Company(
        user_id=company_data.user_id,
        cnpj=company_data.cnpj,
        razao_social=company_data.razao_social,
        nome_fantasia=company_data.nome_fantasia,
        setor=company_data.setor,
        cep=company_data.cep,
        pessoa_de_contato=company_data.pessoa_de_contato,
        fone=company_data.fone,
        site=company_data.site,
        is_active=True,
        is_verified=False
    )
    
    db.add(company)
    db.commit()
    db.refresh(company)
    
    return {
        "id": company.id,
        "user_id": company.user_id,
        "cnpj": company.cnpj,
        "razao_social": company.razao_social,
        "setor": company.setor,
        "is_active": company.is_active,
        "created_at": company.created_at,
        "message": "Empresa criada com sucesso"
    }


@router.get("/empresas")
async def get_all_companies(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(10, ge=1, le=100, description="Número máximo de registros a retornar (máximo 100)"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todas as empresas cadastradas no sistema com paginação
    
    Apenas administradores podem acessar este endpoint.
    
    Query Parameters:
    - skip: Número de registros a pular (default: 0)
    - limit: Número máximo de registros por página (default: 10, máximo: 100)
    """
    total = db.query(Company).count()
    companies = db.query(Company).offset(skip).limit(limit).all()
    
    # Formatar resposta com informações relevantes
    companies_list = []
    for company in companies:
        company_info = {
            "id": company.id,
            "cnpj": company.cnpj,
            "razao_social": company.razao_social,
            "setor": company.setor,
            "cep": company.cep,
            "pessoa_de_contato": company.pessoa_de_contato,
            "fone": company.fone,
            "is_active": company.is_active,
            "is_verified": company.is_verified,
            "created_at": company.created_at,
            "updated_at": company.updated_at,
            "user": {
                "id": company.user.id,
                "email": company.user.email,
                "is_active": company.user.is_active
            } if company.user else None,
            "jobs_count": len(company.jobs) if company.jobs else 0
        }
        companies_list.append(company_info)
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "returned": len(companies_list),
        "empresas": companies_list
    }


@router.get("/candidatos")
async def get_all_candidates(
    skip: int = Query(0, ge=0, description="Número de registros a pular"),
    limit: int = Query(10, ge=1, le=100, description="Número máximo de registros a retornar (máximo 100)"),
    busca: Optional[str] = Query(None, description="Buscar por nome ou email"),
    localizacao: Optional[str] = Query(None, description="Filtrar por estado"),
    habilidade: Optional[str] = Query(None, description="Filtrar por habilidade"),
    experiencia_min: Optional[int] = Query(None, description="Anos de experiência mínima"),
    experiencia_max: Optional[int] = Query(None, description="Anos de experiência máxima"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna candidatos com filtros avançados e paginação
    
    Query Parameters:
    - skip: Número de registros a pular (default: 0)
    - limit: Número máximo de registros por página (default: 10, máximo: 100)
    
    Filtros disponíveis:
    - busca: Nome ou email
    - localizacao: Estado (ex: SP, RJ)
    - habilidade: Habilidade técnica (ex: React, Python)
    - experiencia_min: Anos mínimos
    - experiencia_max: Anos máximos
    """
    query = db.query(Candidate)
    
    # Filtro de busca (nome ou email)
    if busca:
        query = query.filter(
            (Candidate.full_name.ilike(f"%{busca}%")) |
            (Candidate.user.has(User.email.ilike(f"%{busca}%")))
        )
    
    # Filtro de localização
    if localizacao:
        query = query.filter(Candidate.estado == localizacao)
    
    # Filtro de habilidade
    if habilidade:
        query = query.filter(Candidate.habilidades.ilike(f"%{habilidade}%"))
    
    # Contar total antes de aplicar paginação
    total = query.count()
    
    # Aplicar paginação
    candidates = query.offset(skip).limit(limit).all()
    
    # Formatar resposta com informações relevantes
    candidates_list = []
    for candidate in candidates:
        candidate_info = {
            "id": candidate.id,
            "cpf": candidate.cpf,
            "email": candidate.user.email if candidate.user else None,
            "nome": candidate.full_name,
            "localizacao": candidate.estado or "-",
            "telefone": candidate.phone,
            "habilidades": candidate.habilidades or "-",
            "experiencia": candidate.experiencia_profissional or "-",
            "anos_experiencia": candidate.anos_experiencia or 0,
            "area_atuacao": candidate.area_atuacao or "-",
            "data_nascimento": candidate.birth_date,
            "is_active": candidate.user.is_active if candidate.user else None,
            "created_at": candidate.created_at
        }
        candidates_list.append(candidate_info)
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "returned": len(candidates_list),
        "candidatos": candidates_list
    }


@router.get("/candidatos-old")
async def get_all_candidates(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna todos os candidatos cadastrados no sistema
    
    Apenas administradores podem acessar este endpoint.
    Retorna: email, nome, estado
    """
    candidates = db.query(Candidate).all()
    
    # Formatar resposta com informações relevantes
    candidates_list = []
    for candidate in candidates:
        candidate_info = {
            "id": candidate.id,
            "cpf": candidate.cpf,
            "email": candidate.user.email if candidate.user else None,
            "nome": candidate.full_name,
            "localizacao": candidate.estado,
            "telefone": candidate.phone,
            "data_nascimento": candidate.birth_date,
            "is_active": candidate.user.is_active if candidate.user else None,
            "created_at": candidate.created_at
        }
        candidates_list.append(candidate_info)
    
    return {
        "total": len(candidates_list),
        "candidatos": candidates_list
    }


@router.get("/list-admins")
async def list_admins(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Lista todos os administradores cadastrados

    Apenas administradores podem acessar este endpoint.
    Retorna: id, email, status (is_active), verificado, criado em, atualizado em
    """
    admins = db.query(User).filter(User.user_type == UserType.admin).all()

    admins_list = []
    for admin in admins:
        admins_list.append({
            "id": admin.id,
            "email": admin.email,
            "status": "ativo" if admin.is_active else "inativo",
            "is_active": admin.is_active,
            "is_verified": admin.is_verified,
            "criado_em": admin.created_at,
            "atualizado_em": admin.updated_at
        })

    return {
        "total": len(admins_list),
        "administradores": admins_list
    }


# ============================================
# ROTAS PARA GERENCIAMENTO DE TESTES
# ============================================

@router.post("/testes", response_model=TestResponse, status_code=status.HTTP_201_CREATED)
async def create_test(
    test_data: TestCreateRequest,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Cria um novo teste técnico com questões e alternativas baseado em competência e área
    
    **Acesso**: Apenas administradores
    
    **Campos obrigatórios**:
    - competencia_id: ID da competência do banco
    - area: Área de atuação (ex: automacao, eletrica, programacao, etc)
    - nivel: Nível de dificuldade (1-5 ou nome do nível)
    - questions: Lista de questões com alternativas
    
    **Exemplo de payload**:
    ```json
    {
      "competencia_id": 15,
      "area": "automacao",
      "nivel": 3,
      "descricao": "Teste avaliativo de PLCs",
      "questions": [
        {
          "texto_questao": "O que é um PLC?",
          "ordem": 1,
          "alternatives": [
            {"texto": "Controlador Lógico Programável", "is_correct": true, "ordem": 1},
            {"texto": "Programa Linear de Computador", "is_correct": false, "ordem": 2}
          ]
        }
      ]
    }
    ```
    
    O nome do teste é gerado automaticamente como: "[CompetênciaX] - Nível Y"
    """
    try:
        # Validar e obter competência
        competencia = db.query(Competencia).filter(Competencia.id == test_data.competencia_id).first()
        
        if not competencia:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Competência com ID {test_data.competencia_id} não encontrada"
            )
        
        # Validar área
        if competencia.area != test_data.area.lower():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Competência com ID {test_data.competencia_id} não pertence à área '{test_data.area}'. Área correta: '{competencia.area}'"
            )
        
        # Gerar nome do teste automaticamente
        nivel_text = test_data.nivel.value if hasattr(test_data.nivel, 'value') else str(test_data.nivel)
        nome_teste = f"{competencia.nome} - {nivel_text}"
        
        logger.info(f"Admin {current_user.id} criando teste para competência {competencia.id}: {nome_teste}")
        logger.info(f"  Área: {test_data.area}, Nível: {nivel_text}, Questões: {len(test_data.questions)}")
        
        # Criar o teste
        new_test = Test(
            nome=nome_teste,
            habilidade=competencia.nome,
            nivel=test_data.nivel,
            descricao=test_data.descricao or f"Teste técnico para {competencia.nome}",
            created_by=current_user.id
        )
        
        db.add(new_test)
        db.flush()  # Para obter o ID do teste antes de criar questões
        
        # Criar as questões e alternativas
        for idx, question_data in enumerate(test_data.questions):
            logger.info(f"  Questão {idx + 1}: {question_data.texto_questao[:50]}...")
            new_question = Question(
                test_id=new_test.id,
                texto_questao=question_data.texto_questao,
                ordem=question_data.ordem
            )
            db.add(new_question)
            db.flush()  # Para obter o ID da questão antes de criar alternativas
            
            # Criar as alternativas
            for alt_idx, alt_data in enumerate(question_data.alternatives):
                new_alternative = Alternative(
                    question_id=new_question.id,
                    texto=alt_data.texto,
                    is_correct=alt_data.is_correct,
                    ordem=alt_data.ordem
                )
                db.add(new_alternative)
        
        db.commit()
        db.refresh(new_test)
        
        logger.info(f"Teste criado com sucesso: ID={new_test.id}, Nome={nome_teste}")
        return new_test
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Erro ao criar teste: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao criar teste: {str(e)}"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao criar teste: {str(e)}"
        )


@router.get("/testes", response_model=List[TestListItemResponse])
async def list_tests(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Lista todos os testes técnicos cadastrados com suas questões e alternativas
    
    Apenas administradores podem acessar este endpoint.
    Retorna todos os testes formatados para o frontend.
    """
    from sqlalchemy.orm import joinedload
    
    # Carregar todos os testes com questões e alternativas
    tests = db.query(Test).options(
        joinedload(Test.questions).joinedload(Question.alternatives)
    ).all()
    
    # Mapear níveis para números
    nivel_map = {
        TestLevel.iniciante: 1,
        TestLevel.basico: 2,
        TestLevel.intermediario: 3,
        TestLevel.avancado: 4,
        TestLevel.expert: 5
    }
    
    tests_list = []
    for test in tests:
        # Processar questões
        questoes = []
        for question in test.questions:
            # Extrair apenas os textos das alternativas
            alternativas = [alt.texto for alt in sorted(question.alternatives, key=lambda a: a.ordem)]
            
            # Encontrar índice da resposta correta
            resposta_correta = next(
                (i for i, alt in enumerate(sorted(question.alternatives, key=lambda a: a.ordem)) if alt.is_correct),
                0
            )
            
            questoes.append(QuestionListItem(
                id=str(question.id),
                pergunta=question.texto_questao,
                alternativas=alternativas,
                respostaCorreta=resposta_correta,
                nivel=nivel_map.get(test.nivel, 1)
            ))
        
        # Criar objeto de teste formatado
        test_item = TestListItemResponse(
            id=str(test.id),
            nome=test.nome,
            descricao=test.descricao or '',
            nivel=nivel_map.get(test.nivel, 1),
            habilidade=test.habilidade,
            questoes=questoes,
            createdAt=test.created_at,
            createdBy=str(test.created_by)
        )
        tests_list.append(test_item)
    
    return tests_list


@router.get("/testes/{test_id}", response_model=TestResponse)
async def get_test(
    test_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna os detalhes completos de um teste específico
    
    Apenas administradores podem acessar este endpoint.
    Inclui todas as questões e alternativas.
    """
    from sqlalchemy.orm import joinedload
    
    # Carregar teste com todas as relações (questões e alternativas)
    test = db.query(Test).options(
        joinedload(Test.questions).joinedload(Question.alternatives)
    ).filter(Test.id == test_id).first()
    
    if not test:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teste não encontrado"
        )
    
    print(f"DEBUG - Retornando teste {test_id}: {len(test.questions)} questões")
    for q in test.questions:
        print(f"  Questão {q.id}: {len(q.alternatives)} alternativas")
    
    return test


@router.put("/testes/{test_id}", response_model=TestResponse)
async def update_test(
    test_id: int,
    test_data: TestUpdate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Atualiza um teste técnico existente
    
    Apenas administradores podem editar testes.
    
    Esta operação substitui completamente:
    - Informações básicas do teste (nome, habilidade, nível, descrição)
    - Todas as questões e alternativas (as antigas são removidas e novas são criadas)
    
    Envie TODAS as questões que deseja manter, incluindo as não modificadas.
    """
    try:
        # Log dos dados recebidos para debug
        print(f"DEBUG - Atualizando teste {test_id}")
        print(f"DEBUG - Dados recebidos: nome={test_data.nome}, nivel={test_data.nivel}, questões={len(test_data.questions)}")
        
        # Buscar o teste existente
        test = db.query(Test).filter(Test.id == test_id).first()
        
        if not test:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teste não encontrado"
            )
        
        # Atualizar informações básicas do teste
        test.nome = test_data.nome
        test.habilidade = test_data.habilidade
        test.nivel = test_data.nivel
        test.descricao = test_data.descricao
        
        # Primeiro, buscar todas as questões do teste
        questions = db.query(Question).filter(Question.test_id == test_id).all()
        
        # Deletar todas as alternativas de cada questão primeiro
        for question in questions:
            db.query(Alternative).filter(Alternative.question_id == question.id).delete()
        
        # Agora deletar as questões
        db.query(Question).filter(Question.test_id == test_id).delete()
        
        # Criar as novas questões e alternativas
        for idx, question_data in enumerate(test_data.questions):
            print(f"DEBUG - Criando questão {idx + 1}: {question_data.texto_questao}")
            new_question = Question(
                test_id=test.id,
                texto_questao=question_data.texto_questao,
                ordem=question_data.ordem
            )
            db.add(new_question)
            db.flush()  # Para obter o ID da questão antes de criar alternativas
            
            # Criar as alternativas
            for alt_idx, alt_data in enumerate(question_data.alternatives):
                print(f"  DEBUG - Alternativa {alt_idx}: {alt_data.texto[:30]}... (correta: {alt_data.is_correct})")
                new_alternative = Alternative(
                    question_id=new_question.id,
                    texto=alt_data.texto,
                    is_correct=alt_data.is_correct,
                    ordem=alt_data.ordem
                )
                db.add(new_alternative)
    
        db.commit()
        db.refresh(test)
        
        print(f"DEBUG - Teste {test_id} atualizado com sucesso")
        return test
    
    except Exception as e:
        db.rollback()
        print(f"ERRO ao atualizar teste {test_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao atualizar teste: {str(e)}"
        )


@router.delete("/testes/{test_id}", status_code=status.HTTP_200_OK)
async def delete_test(
    test_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Deleta um teste técnico
    
    Apenas administradores podem deletar testes.
    Remove o teste e todas as questões e alternativas associadas (cascade).
    """
    from sqlalchemy.orm import joinedload
    
    # Carregar teste com relações para contar antes de deletar
    test = db.query(Test).options(
        joinedload(Test.questions).joinedload(Question.alternatives)
    ).filter(Test.id == test_id).first()
    
    if not test:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teste não encontrado"
        )
    
    # Contar para log
    total_questions = len(test.questions)
    total_alternatives = sum(len(q.alternatives) for q in test.questions)
    
    print(f"DEBUG - Deletando teste {test_id}: {total_questions} questões, {total_alternatives} alternativas")
    
    # Delete irá remover questões e alternativas via cascade
    db.delete(test)
    db.commit()
    
    print(f"DEBUG - Teste {test_id} deletado com sucesso")
    
    return {
        "message": "Teste deletado com sucesso",
        "test_id": test_id,
        "questoes_removidas": total_questions,
        "alternativas_removidas": total_alternatives
    }


# ============================================================================
# ENDPOINTS DE IMPORTAÇÃO DE QUESTÕES
# ============================================================================

class ImportTestsResponse(BaseModel):
    """Response da importação de testes"""
    sucesso: bool
    total_linhas_processadas: int
    testes_criados: int
    questoes_criadas: int
    alternativas_criadas: int
    erros: List[str]
    avisos: List[str]
    
    class Config:
        from_attributes = True


@router.post("/testes/importar", response_model=ImportTestsResponse)
async def importar_testes_arquivo(
    arquivo: UploadFile = File(...),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Importa questões de testes a partir de arquivo Excel (.xlsx) ou CSV (.csv)
    
    **Formato do arquivo esperado:**
    
    | habilidade | nivel | pergunta | opcao_a | opcao_b | opcao_c | opcao_d | resposta_correta |
    |---|---|---|---|---|---|---|---|
    | Python | Básico | O que é uma lista? | Estrutura mutável | String | Número | Tupla | A |
    | Python | Intermediário | O que é decorator? | Função que modifica | Comentário | Variável | Loop | A |
    | React | Avançado | O que é Context API? | State management | Hook | Componente | Props | A |
    
    **Colunas obrigatórias:**
    - `habilidade`: Nome da habilidade (Python, React, JavaScript, etc)
    - `nivel`: Um de: Básico, Intermediário, Avançado
    - `pergunta`: Texto da questão
    - `opcao_a`: Texto da alternativa A
    - `opcao_b`: Texto da alternativa B
    - `opcao_c`: Texto da alternativa C
    - `opcao_d`: Texto da alternativa D
    - `resposta_correta`: A, B, C ou D (qual é a correta)
    
    **Exemplos de Nível:**
    - Básico (ou basico)
    - Intermediário (ou intermediario)
    - Avançado (ou avancado)
    
    **Formato do arquivo:**
    - Excel: .xlsx com dados em primeira planilha
    - CSV: UTF-8, separado por vírgula
    
    **Response:**
    - `sucesso`: Se a importação foi bem-sucedida
    - `questoes_criadas`: Quantas questões foram criadas
    - `erros`: Lista de erros encontrados (linhas inválidas)
    - `avisos`: Avisos (ex: teste já existente)
    """
    
    try:
        # Validar extensão do arquivo
        if not arquivo.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Arquivo sem nome"
            )
        
        extensao = arquivo.filename.lower().split('.')[-1]
        
        if extensao not in ['xlsx', 'csv']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Arquivo deve ser .xlsx ou .csv"
            )
        
        # Ler conteúdo do arquivo
        conteudo = await arquivo.read()
        
        if not conteudo:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Arquivo está vazio"
            )
        
        # Processar arquivo
        if extensao == 'xlsx':
            dados = TestImportService.ler_excel(conteudo)
        else:  # csv
            dados = TestImportService.ler_csv(conteudo)
        
        if not dados:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Nenhuma linha de dados encontrada no arquivo"
            )
        
        # Importar questões
        stats = TestImportService.importar_questoes(
            dados=dados,
            db=db,
            admin_user_id=current_user.id
        )
        
        sucesso = len(stats["erros"]) == 0
        
        return ImportTestsResponse(
            sucesso=sucesso,
            total_linhas_processadas=stats["total_linhas"],
            testes_criados=stats["testes_criados"],
            questoes_criadas=stats["questoes_criadas"],
            alternativas_criadas=stats["alternativas_criadas"],
            erros=stats["erros"],
            avisos=stats["avisos"]
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao processar arquivo: {str(e)}"
        )


@router.get("/testes/template-excel")
async def baixar_template_excel(current_user: User = Depends(get_current_admin)):
    """
    Retorna um arquivo Excel de template para importar questões
    
    Use este template como guia para estruturar seu arquivo.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import FileResponse
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questões"
        
        # Headers
        headers = [
            "habilidade",
            "nivel",
            "pergunta",
            "opcao_a",
            "opcao_b",
            "opcao_c",
            "opcao_d",
            "resposta_correta"
        ]
        
        ws.append(headers)
        
        # Estilizar header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Adicionar exemplos
        exemplos = [
            ["Python", "Básico", "O que é uma lista em Python?", "Estrutura mutável que armazena elementos", "Uma string imutável", "Um número inteiro", "Um tipo como tupla", "A"],
            ["Python", "Básico", "Como declarar uma função?", "def nome_funcao():", "function nome_funcao()", "func nome_funcao()", "void nome_funcao()", "A"],
            ["Python", "Intermediário", "O que são decorators?", "Funções que modificam outras funções", "Comentários especiais", "Tipos de variáveis", "Estruturas de controle", "A"],
            ["React", "Básico", "O que é JSX?", "Sintaxe que permite HTML em JavaScript", "Uma biblioteca CSS", "Um tipo de arquivo", "Uma ferramenta de build", "A"],
            ["JavaScript", "Avançado", "O que é Event Loop?", "Mecanismo que executa callbacks em JavaScript", "Um tipo de loop", "Uma função", "Um tipo de erro", "A"],
        ]
        
        for exemplo in exemplos:
            ws.append(exemplo)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15
        
        # Adicionar folha com instruções
        ws_info = wb.create_sheet("Instruções")
        
        instrucoes = [
            ["GUIA DE IMPORTAÇÃO DE QUESTÕES", ""],
            ["", ""],
            ["Colunas Obrigatórias:", ""],
            ["- habilidade", "Nome da habilidade: Python, React, JavaScript, etc"],
            ["- nivel", "Um de: Básico, Intermediário, Avançado (ou basico, intermediario, avancado)"],
            ["- pergunta", "Texto da questão"],
            ["- opcao_a", "Primeira opção de resposta"],
            ["- opcao_b", "Segunda opção de resposta"],
            ["- opcao_c", "Terceira opção de resposta"],
            ["- opcao_d", "Quarta opção de resposta"],
            ["- resposta_correta", "A, B, C ou D (qual é a resposta correta)"],
            ["", ""],
            ["Dicas:", ""],
            ["1. Não deixe células vazias", ""],
            ["2. Use UTF-8 se for salvar como CSV", ""],
            ["3. A primeira linha deve ter os nomes das colunas", ""],
            ["4. Use nomes exatos das colunas (maiúsculas/minúsculas não importam)", ""],
            ["5. Nível pode ser: Básico, Intermediário ou Avançado", ""],
            ["", ""],
            ["Exemplo de linha:", ""],
            ["Python | Básico | O que é uma lista? | Estrutura mutável | String | Número | Tupla | A"],
        ]
        
        for instrucao in instrucoes:
            ws_info.append(instrucao)
        
        ws_info.column_dimensions['A'].width = 40
        ws_info.column_dimensions['B'].width = 50
        
        # Salvar em memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_questoes.xlsx"}
        )
    
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="openpyxl não instalado. Instale com: pip install openpyxl"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao gerar template: {str(e)}"
        )


# ============================================================================
# ENDPOINTS DE DETALHES ESPECÍFICOS PARA MODAIS
# ============================================================================

@router.get("/candidatos/{candidato_id}")
async def get_candidato_details(
    candidato_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna detalhes completos de um candidato específico para modal
    """
    candidate = db.query(Candidate).filter(Candidate.id == candidato_id).first()
    
    if not candidate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Candidato não encontrado"
        )
    
    return {
        "id": candidate.id,
        "cpf": candidate.cpf,
        "full_name": candidate.full_name,
        "email": candidate.user.email if candidate.user else None,
        "phone": candidate.phone,
        "birth_date": candidate.birth_date,
        "genero": candidate.genero,
        "estado_civil": candidate.estado_civil,
        "location": candidate.location,
        "estado": candidate.estado,
        "cidade": candidate.cidade,
        "cep": candidate.cep,
        "logradouro": candidate.logradouro,
        "numero": candidate.numero,
        "complemento": candidate.complemento,
        "bairro": candidate.bairro,
        "is_pcd": candidate.is_pcd,
        "tipo_pcd": candidate.tipo_pcd,
        "necessidades_adaptacao": candidate.necessidades_adaptacao,
        "experiencia_profissional": candidate.experiencia_profissional,
        "anos_experiencia": candidate.anos_experiencia,
        "formacao_escolaridade": candidate.formacao_escolaridade,
        "habilidades": candidate.habilidades,
        "area_atuacao": candidate.area_atuacao,
        "resume_url": candidate.resume_url,
        "linkedin_url": candidate.linkedin_url,
        "portfolio_url": candidate.portfolio_url,
        "bio": candidate.bio,
        "is_active": candidate.is_active,
        "onboarding_completo": candidate.onboarding_completo,
        "created_at": candidate.created_at,
        "updated_at": candidate.updated_at
    }


@router.get("/usuarios/{usuario_id}")
async def get_usuario_details(
    usuario_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna detalhes completos de um usuário específico para modal
    """
    user = db.query(User).filter(User.id == usuario_id).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuário não encontrado"
        )
    
    result = {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "user_type": user.user_type.value,
        "is_active": user.is_active,
        "created_at": user.created_at,
        "updated_at": user.updated_at
    }
    
    # Adicionar detalhes específicos por tipo de usuário
    if user.user_type == UserType.candidato and user.candidate:
        result["candidate_details"] = {
            "id": user.candidate.id,
            "cpf": user.candidate.cpf,
            "full_name": user.candidate.full_name,
            "phone": user.candidate.phone,
            "anos_experiencia": user.candidate.anos_experiencia,
            "area_atuacao": user.candidate.area_atuacao,
            "habilidades": user.candidate.habilidades
        }
    elif user.user_type == UserType.empresa and user.company:
        result["company_details"] = {
            "id": user.company.id,
            "cnpj": user.company.cnpj,
            "razao_social": user.company.razao_social,
            "email": user.company.email,
            "phone": user.company.phone
        }
    
    return result


@router.get("/candidaturas/{candidatura_id}")
async def get_candidatura_details(
    candidatura_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna detalhes completos de uma candidatura específica para modal
    """
    app = db.query(JobApplication).filter(JobApplication.id == candidatura_id).first()
    
    if not app:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Candidatura não encontrada"
        )
    
    return {
        "id": app.id,
        "status": app.status.value,
        "created_at": app.created_at,
        "updated_at": app.updated_at,
        "cover_letter": app.cover_letter,
        "screening_answers": app.screening_answers,
        "job": {
            "id": app.job.id,
            "title": app.job.title,
            "description": app.job.description,
            "location": app.job.location,
            "job_type": app.job.job_type,
            "salary_min": float(app.job.salary_min) if app.job.salary_min else None,
            "salary_max": float(app.job.salary_max) if app.job.salary_max else None,
            "status": app.job.status.value,
            "company": {
                "id": app.job.company.id,
                "cnpj": app.job.company.cnpj,
                "razao_social": app.job.company.razao_social
            }
        } if app.job else None,
        "candidate": {
            "id": app.candidate.id,
            "cpf": app.candidate.cpf,
            "full_name": app.candidate.full_name,
            "phone": app.candidate.phone,
            "email": app.candidate.user.email if app.candidate.user else None,
            "habilidades": app.candidate.habilidades,
            "experiencia_profissional": app.candidate.experiencia_profissional,
            "anos_experiencia": app.candidate.anos_experiencia,
            "area_atuacao": app.candidate.area_atuacao
        } if app.candidate else None
    }


@router.get("/empresas/{empresa_id}")
async def get_empresa_details(
    empresa_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Retorna detalhes completos de uma empresa específica para modal
    """
    company = db.query(Company).filter(Company.id == empresa_id).first()
    
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa não encontrada"
        )
    
    return {
        "id": company.id,
        "cnpj": company.cnpj,
        "razao_social": company.razao_social,
        "email": company.email,
        "phone": company.phone,
        "website": company.website,
        "logo_url": company.logo_url,
        "description": company.description,
        "is_active": company.is_active,
        "created_at": company.created_at,
        "updated_at": company.updated_at
    }


# ============================================================================
# GERENCIAMENTO DE COMPETÊNCIAS
# ============================================================================

@router.post("/competencias", response_model=CompetenciaResponse, status_code=status.HTTP_201_CREATED)
async def criar_competencia(
    dados: CompetenciaCreate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Cria uma nova competência que será exibida para candidatos
    
    **Acesso**: Apenas administradores
    
    **Exemplo de payload**:
    ```json
    {
        "nome": "PLCs (Controladores Lógicos Programáveis)",
        "descricao": "Conhecimento em programação de controladores lógicos programáveis",
        "categoria": "tecnica",
        "area": "automacao"
    }
    ```
    
    **Áreas disponíveis**:
    - eletrica
    - manutencao
    - automacao
    - mecanica
    - civil
    - pneumatica
    - hidraulica
    - eletronica
    - programacao
    - redes
    """
    logger.info(f"Admin {current_user.id} criando nova competência: {dados.nome}")
    
    # Validar se área é válida
    try:
        area_enum = AreaAtuacao(dados.area.lower())
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Área inválida: {dados.area}. Áreas válidas: {', '.join([a.value for a in AreaAtuacao])}"
        )
    
    # Verificar se competência já existe
    existente = db.query(Competencia).filter(
        Competencia.nome == dados.nome,
        Competencia.area == area_enum.value
    ).first()
    
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Competência '{dados.nome}' já existe na área '{dados.area}'"
        )
    
    # Criar nova competência
    nova_competencia = Competencia(
        nome=dados.nome,
        descricao=dados.descricao,
        categoria=dados.categoria or "tecnica",
        area=area_enum.value
    )
    
    db.add(nova_competencia)
    db.commit()
    db.refresh(nova_competencia)
    
    logger.info(f"Competência criada com sucesso: {nova_competencia.id}")
    
    return CompetenciaResponse(
        id=nova_competencia.id,
        nome=nova_competencia.nome,
        descricao=nova_competencia.descricao,
        categoria=nova_competencia.categoria,
        area=AreaAtuacao(nova_competencia.area),
        created_at=nova_competencia.created_at
    )


@router.get("/competencias", response_model=List[CompetenciaResponse])
async def listar_competencias(
    area: Optional[str] = Query(None, description="Filtrar por área"),
    categoria: Optional[str] = Query(None, description="Filtrar por categoria"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Lista todas as competências cadastradas
    
    **Acesso**: Apenas administradores
    
    **Filtros disponíveis**:
    - area: Filtra por área de atuação (ex: automacao, eletrica)
    - categoria: Filtra por categoria (ex: tecnica, soft)
    """
    logger.info(f"Admin {current_user.id} listando competências")
    
    query = db.query(Competencia)
    
    if area:
        try:
            area_enum = AreaAtuacao(area.lower())
            query = query.filter(Competencia.area == area_enum.value)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Área inválida: {area}"
            )
    
    if categoria:
        query = query.filter(Competencia.categoria == categoria)
    
    competencias = query.all()
    
    result = []
    for c in competencias:
        try:
            area_enum = AreaAtuacao(c.area)
        except ValueError:
            # Se a área armazenada não é válida, pula essa competência
            logger.warning(f"Competência {c.id} tem área inválida: {c.area}")
            continue
        
        result.append(CompetenciaResponse(
            id=c.id,
            nome=c.nome,
            descricao=c.descricao,
            categoria=c.categoria,
            area=area_enum,
            created_at=c.created_at
        ))
    
    return result


@router.get("/competencias/{competencia_id}", response_model=CompetenciaResponse)
async def obter_competencia(
    competencia_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Obtém os detalhes de uma competência específica
    
    **Acesso**: Apenas administradores
    """
    logger.info(f"Admin {current_user.id} obtendo competência {competencia_id}")
    
    competencia = db.query(Competencia).filter(Competencia.id == competencia_id).first()
    
    if not competencia:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Competência não encontrada"
        )
    
    return CompetenciaResponse(
        id=competencia.id,
        nome=competencia.nome,
        descricao=competencia.descricao,
        categoria=competencia.categoria,
        area=AreaAtuacao(competencia.area),
        created_at=competencia.created_at
    )


@router.put("/competencias/{competencia_id}", response_model=CompetenciaResponse)
async def atualizar_competencia(
    competencia_id: int,
    dados: CompetenciaCreate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Atualiza uma competência existente
    
    **Acesso**: Apenas administradores
    """
    logger.info(f"Admin {current_user.id} atualizando competência {competencia_id}")
    
    competencia = db.query(Competencia).filter(Competencia.id == competencia_id).first()
    
    if not competencia:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Competência não encontrada"
        )
    
    # Validar área se foi alterada
    if dados.area:
        try:
            area_enum = AreaAtuacao(dados.area.lower())
            competencia.area = area_enum.value
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Área inválida: {dados.area}"
            )
    
    competencia.nome = dados.nome
    competencia.descricao = dados.descricao
    competencia.categoria = dados.categoria or "tecnica"
    
    db.commit()
    db.refresh(competencia)
    
    logger.info(f"Competência {competencia_id} atualizada com sucesso")
    
    return CompetenciaResponse(
        id=competencia.id,
        nome=competencia.nome,
        descricao=competencia.descricao,
        categoria=competencia.categoria,
        area=AreaAtuacao(competencia.area),
        created_at=competencia.created_at
    )


@router.delete("/competencias/{competencia_id}", status_code=status.HTTP_204_NO_CONTENT)
async def deletar_competencia(
    competencia_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Deleta uma competência
    
    **Acesso**: Apenas administradores
    
    **Nota**: Competências que estão sendo usadas em autoavaliações serão deletadas em cascata
    """
    logger.info(f"Admin {current_user.id} deletando competência {competencia_id}")
    
    competencia = db.query(Competencia).filter(Competencia.id == competencia_id).first()
    
    if not competencia:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Competência não encontrada"
        )
    
    # Contar quantas autoavaliações estão usando essa competência
    autoavaliacoes_count = len(competencia.autoavaliacoes) if competencia.autoavaliacoes else 0
    
    if autoavaliacoes_count > 0:
        logger.warning(f"Deletando competência {competencia_id} que está em uso em {autoavaliacoes_count} autoavaliações")
    
    db.delete(competencia)
    db.commit()
    
    logger.info(f"Competência {competencia_id} deletada com sucesso")


@router.get("/competencias-por-area/{area}", response_model=List[CompetenciaResponse])
async def listar_competencias_por_area(
    area: str,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Lista todas as competências de uma área específica
    
    **Acesso**: Apenas administradores
    
    **Áreas disponíveis**:
    - eletrica
    - manutencao
    - automacao
    - mecanica
    - civil
    - pneumatica
    - hidraulica
    - eletronica
    - programacao
    - redes
    """
    logger.info(f"Admin {current_user.id} listando competências da área: {area}")
    
    try:
        area_enum = AreaAtuacao(area.lower())
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Área inválida: {area}. Áreas válidas: {', '.join([a.value for a in AreaAtuacao])}"
        )
    
    competencias = db.query(Competencia).filter(Competencia.area == area_enum.value).all()
    
    return [
        CompetenciaResponse(
            id=c.id,
            nome=c.nome,
            descricao=c.descricao,
            categoria=c.categoria,
            area=AreaAtuacao(c.area),
            created_at=c.created_at
        )
        for c in competencias
    ]
